import json
import os
import sys
from difflib import SequenceMatcher

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side


FILES = {
    "colectiva": "Control de temperaturas bombas colectivas.xlsx",
    "selectiva": "Control de temperatura bombas selectiva.xlsx",
}

BORDER = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)


def closest_sheet(tag, sheetnames):
    target = tag.strip().upper()
    best = None
    best_ratio = 0.0
    for name in sheetnames:
        ratio = SequenceMatcher(None, target, name.strip().upper()).ratio()
        if ratio > best_ratio:
            best = name
            best_ratio = ratio
    return best if best_ratio > 0.8 else None


def normalise_number(value):
    text = str(value or "").strip().replace(",", ".")
    if not text:
        return ""
    if text == "Fuera de servicio":
        return text
    try:
        number = float(text)
    except ValueError:
        return text
    return str(number) if 0 <= number <= 100 else ""


def ensure_workbook(path_name):
    if os.path.exists(path_name):
        return load_workbook(path_name)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def ensure_sheet(wb, plant, tag):
    if tag in wb.sheetnames:
        return wb[tag]
    match = closest_sheet(tag, wb.sheetnames)
    if match:
        return wb[match]

    ws = wb.create_sheet(title=tag)
    if plant == "colectiva":
        headers = ["TAG", "Ubicación", "Fecha", "Lado Bomba", "Lado Polea", "Lado bomba", "Lado Polea", "Sin acceso"]
    else:
        headers = ["TAG", "Ubicación", "Fecha", "Lado Libre", "Lado Polea", "Lado bomba", "Lado Polea", "Sin acceso"]
    for col, value in enumerate(headers, start=2):
        cell = ws.cell(row=1, column=col, value=value)
        cell.border = BORDER
    return ws


def write_record(ws, row_data, excel_date):
    row_to_update = None
    tag = row_data[0]
    for row_num in range(4, ws.max_row + 1):
        if ws.cell(row=row_num, column=2).value == tag and ws.cell(row=row_num, column=4).value == excel_date:
            row_to_update = row_num
            break

    if row_to_update is None:
        row_to_update = 4
        while row_to_update <= ws.max_row:
            if ws.cell(row=row_to_update, column=2).value is None and ws.cell(row=row_to_update, column=4).value is None:
                break
            row_to_update += 1
        if row_to_update > ws.max_row:
            row_to_update = ws.max_row + 1

    for col, value in enumerate(row_data, start=2):
        cell = ws.cell(row=row_to_update, column=col)
        if cell.coordinate in ws.merged_cells:
            continue
        cell.value = value
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left" if col <= 3 else "right")


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Payload requerido")
    with open(sys.argv[1], "r", encoding="utf-8") as fh:
        payload = json.load(fh)

    plant = payload["plant"]
    file_name = FILES[plant]
    workbook_path = os.path.join(os.getcwd(), "bomba16", file_name)
    wb = ensure_workbook(workbook_path)
    day, month, year = payload["date"].split("/")
    excel_date = f"{day}-{month}-{year.zfill(4)[-2:]}"

    for record in payload["records"]:
        tag = record["tag"]
        ws = ensure_sheet(wb, plant, tag)
        row_data = [
            tag,
            record["ubicacion"],
            excel_date,
            normalise_number(record.get("v1")),
            normalise_number(record.get("v2")),
            normalise_number(record.get("v3")),
            normalise_number(record.get("v4")),
            record.get("observacion") or "",
        ]
        write_record(ws, row_data, excel_date)

    wb.save(workbook_path)
    print(json.dumps({"file": file_name}, ensure_ascii=True))


if __name__ == "__main__":
    main()
