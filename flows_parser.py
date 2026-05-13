import json
import sys

from openpyxl import load_workbook


def main():
    if len(sys.argv) != 2:
        raise SystemExit("usage: flows_parser.py <flujos.xlsx>")
    wb = load_workbook(sys.argv[1], data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(json.dumps({"tags": [], "flows": {}}))
        return

    header = rows[0]
    tags = []
    flows = {}
    for col in range(1, len(header)):
        tag_key = str(header[col]).strip().lower() if header[col] else ""
        tag_display = str(header[col]).strip().upper() if header[col] else ""
        if not tag_key:
            continue
        tag_flows = []
        for r in range(1, len(rows)):
            val = rows[r][col] if col < len(rows[r]) else None
            if val and str(val).strip():
                tag_flows.append(str(val).strip())
        tags.append({"key": tag_key, "display": tag_display})
        flows[tag_key] = tag_flows

    result = {"tags": tags, "flows": flows}
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()