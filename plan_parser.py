import json
import sys
from datetime import datetime, timedelta

from openpyxl import load_workbook


SECTIONS = {
    "colectiva": (4, 105),
    "selectiva": (108, 164),
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def parse_date(value, week_start, base_date):
    if isinstance(value, datetime):
        return (week_start + timedelta(days=(value.date() - base_date).days)).isoformat()
    if " AL " in clean(value).upper():
        week_end = week_start + timedelta(days=6)
        return f"{week_start.strftime('%d/%m/%y')} AL {week_end.strftime('%d/%m/%y')}"
    return clean(value)


def normalize_status(value):
    text = clean(value).lower()
    if text in {"", "pendiente", "pending"}:
        return "Pendiente"
    if text in {"cancelada", "cancelado"}:
        return "Cancelada"
    if text in {"en proceso", "proceso"}:
        return "En proceso"
    if text == "ok":
        return "Ok"
    return "Pendiente"


def main():
    if len(sys.argv) != 3:
        raise SystemExit("usage: plan_parser.py <plan.xlsx> <week-start-yyyy-mm-dd>")
    workbook_path = sys.argv[1]
    week_start = datetime.fromisoformat(sys.argv[2]).date()
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["Plan Turno Sist Analisis."]

    base_dates = []
    for start, end in SECTIONS.values():
        for row in ws.iter_rows(min_row=start, max_row=end, max_col=9, values_only=True):
            if isinstance(row[0], datetime):
                base_dates.append(row[0].date())
    base_date = min(base_dates)

    result = {"colectiva": [], "selectiva": []}
    for plant, (start, end) in SECTIONS.items():
        for row_number, row in enumerate(ws.iter_rows(min_row=start, max_row=end, max_col=9, values_only=True), start=start):
            if not any(value is not None for value in row):
                continue
            if not isinstance(row[0], datetime) and " AL " not in clean(row[0]).upper():
                continue
            day = parse_date(row[0], week_start, base_date)
            if not day:
                continue
            result[plant].append(
                {
                    "rowKey": f"{plant}-{row_number}",
                    "day": day,
                    "turn": clean(row[3]),
                    "workGroup": clean(row[4]),
                    "status": normalize_status(row[5]),
                    "workOrder": clean(row[6]),
                    "equipment": clean(row[7]),
                    "description": clean(row[8]),
                }
            )
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
