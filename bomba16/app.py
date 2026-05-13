import os
from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import sqlite3
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from difflib import SequenceMatcher
import json

os.makedirs('instance', exist_ok=True)

app = Flask(__name__)
app.secret_key = 'clave_bombas_final'

# Archivos predeterminados (no se cambian dinámicamente desde la UI)
ARCHIVO_COLECTIVA = "Control de temperaturas bombas colectivas.xlsx"
ARCHIVO_SELECTIVA = "Control de temperatura bombas selectiva.xlsx"

# --- DEFINICIÓN DIRECTA DE BOMBAS ---
BOMBAS = {
    'colectiva': [
        {"tag": "3311-PP-006", "ubicacion": "Relave Scavenger"},
        {"tag": "3311-PP-007", "ubicacion": "Colas Fila 1"},
        {"tag": "3311-PP-008", "ubicacion": "Colas Fila 2"},
        {"tag": "3311-PP-009", "ubicacion": "Colas Fila 3"},
        {"tag": "3511-PP-012", "ubicacion": "Relave 1ra Limpieza"},
        {"tag": "3311-PP-013", "ubicacion": "Concentrado Fila 1"},
        {"tag": "3311-PP-014", "ubicacion": "Concentrado Fila 2"},
        {"tag": "3311-PP-015", "ubicacion": "Concentrado Fila 3"},
        {"tag": "3511-PP-018", "ubicacion": "Concentrado Scavenger"},
        {"tag": "3511-PP-019", "ubicacion": "Concentrado 1ra Limpieza"},
        {"tag": "3511-PP-023", "ubicacion": "Rechazo Courier Remolienda"},
    ],
    'selectiva': [
        {"tag": "3811-PP-620", "ubicacion": "Concentrado primario Fila 2"},
        {"tag": "3811-PP-621", "ubicacion": "Concentrado primario Fila 1"},
        {"tag": "3811-PP-622", "ubicacion": "Cola primaria fila 2"},
        {"tag": "3811-PP-623", "ubicacion": "Cola primaria fila 1"},
        {"tag": "3811-PP-624", "ubicacion": "Concentrado pre primario"},
        {"tag": "3811-PP-625", "ubicacion": "Concentrado colectivo"},
        {"tag": "3811-PP-626", "ubicacion": "Concentrado primera limpieza"},
        {"tag": "3811-PP-627", "ubicacion": "Concentrado segunda limpieza"},
        {"tag": "3811-PP-628", "ubicacion": "Concentrado tercera limpieza"},
        {"tag": "3811-PP-637", "ubicacion": "Rechazo BX632"},
        {"tag": "3811-PP-638", "ubicacion": "Rechazo BX633"},
        {"tag": "3811-PP-641", "ubicacion": "Rechazo BX634"},
    ]
}

# --- BASE DE DATOS ---
def init_db():
    """Inicializa la base de datos con índice en la columna fecha"""
    conn = sqlite3.connect('instance/temperaturas.db')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            planta TEXT NOT NULL,
            tag TEXT NOT NULL,
            ubicacion TEXT NOT NULL,
            fecha TEXT NOT NULL, -- Formato dd/mm/yy
            v1 TEXT, v2 TEXT, v3 TEXT, v4 TEXT,
            observacion TEXT
        )
    ''')
    # Crear índice para acelerar búsquedas por fecha
    conn.execute('CREATE INDEX IF NOT EXISTS idx_fecha ON registros (fecha)')
    conn.commit()
    conn.close()

init_db()

# --- FUNCIONES DE APOYO ---
def es_valido_temp(valor, permitir_fos=False):
    """Valida que un valor sea una temperatura válida o 'Fuera de servicio'"""
    if valor == "": return True
    if permitir_fos and valor == "Fuera de servicio": return True
    try:
        f = float(valor.replace(',', '.'))
        return 0 <= f <= 100
    except:
        return False

def normalizar_numero(valor):
    """Convierte coma decimal a punto y valida rango"""
    if valor == "": return ""
    valor = valor.replace(',', '.')
    try:
        f = float(valor)
        if 0 <= f <= 100:
            return str(f)
        else:
            return None
    except:
        return None

def encontrar_hoja_mas_parecida(tag_objetivo, nombres_hojas):
    """Encuentra la hoja con nombre más parecido al TAG"""
    tag_objetivo = tag_objetivo.strip().upper()
    mejor_coincidencia = None
    mejor_ratio = 0.0
    for nombre in nombres_hojas:
        nombre_limpio = nombre.strip().upper()
        ratio = SequenceMatcher(None, tag_objetivo, nombre_limpio).ratio()
        if ratio > mejor_ratio:
            mejor_ratio = ratio
            mejor_coincidencia = nombre
    return mejor_coincidencia if mejor_ratio > 0.8 else None

# --- Definición de borde más sutil ---
borde_sutil = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# --- RUTAS ---
@app.route('/')
def index():
    """Página principal: selector de fechas existentes y todas las bombas"""
    try:
        conn = sqlite3.connect('instance/temperaturas.db')
        cur = conn.cursor()
        # Obtener fechas únicas ordenadas DESC - CORREGIDO PARA ORDEN CORRECTO
        # Convertimos dd/mm/yy a yy-mm-dd para ordenar correctamente
        cur.execute("""
            SELECT DISTINCT fecha 
            FROM registros 
            ORDER BY 
                substr(fecha, 7, 2) || '-' || 
                substr(fecha, 4, 2) || '-' || 
                substr(fecha, 1, 2) DESC
        """)
        fechas_tuplas = cur.fetchall()
        conn.close()
        
        fechas_disponibles = [f[0] for f in fechas_tuplas]
        fecha_hoy = datetime.now().strftime('%d/%m/%y')
        
        return render_template('index.html', 
                             bombas=BOMBAS, 
                             fecha_default=fecha_hoy, 
                             fechas_disponibles=fechas_disponibles)
    except Exception as e:
        print(f"Error al cargar fechas: {e}")
        return render_template('index.html', 
                             bombas=BOMBAS, 
                             fecha_default=datetime.now().strftime('%d/%m/%y'), 
                             fechas_disponibles=[])

@app.route('/cargar', methods=['POST'])
def cargar():
    """Cargar datos para una fecha específica, mostrando siempre todas las bombas"""
    dia = request.form.get('dia', '').strip()
    mes = request.form.get('mes', '').strip()
    ano = request.form.get('ano', '').strip()
    if not all([dia, mes, ano]):
        return jsonify({"error": "Fecha incompleta"}), 400
    fecha = f"{dia}/{mes}/{ano}"
    conn = sqlite3.connect('instance/temperaturas.db')
    cur = conn.cursor()
    cur.execute("SELECT planta, tag, v1, v2, v3, v4, observacion FROM registros WHERE fecha = ?", (fecha,))
    filas = cur.fetchall()
    conn.close()
    
    # Inicializamos datos para ambas plantas con entradas vacías para cada bomba
    datos = {'colectiva': {}, 'selectiva': {}}
    for planta_tipo in ['colectiva', 'selectiva']:
        for bomba in BOMBAS[planta_tipo]:
            tag = bomba['tag']
            datos[planta_tipo][tag] = ["", "", "", "", ""] # Inicializar con 5 valores vacíos (v1-v4, obs)

    # Luego sobrescribimos con los datos encontrados en la base de datos
    for planta, tag, v1, v2, v3, v4, obs in filas:
        if planta in datos and tag in datos[planta]:
            datos[planta][tag] = [v1 or "", v2 or "", v3 or "", v4 or "", obs or ""]
    return jsonify(datos)

@app.route('/guardar', methods=['POST'])
def guardar():
    """Guardar los datos de una planta"""
    planta = request.form.get('planta')
    dia = request.form.get('dia', '').strip()
    mes = request.form.get('mes', '').strip()
    ano = request.form.get('ano', '').strip()
    if not all([dia, mes, ano]) or planta not in BOMBAS:
        return jsonify({"error": "Falta fecha o planta inválida"}), 400
    fecha = f"{dia}/{mes}/{ano}"
    bombas = BOMBAS[planta]
    try:
        conn = sqlite3.connect('instance/temperaturas.db')
        for i, b in enumerate(bombas):
            sin_acceso = request.form.get(f'sin_acceso_{planta}_{i}', '') == 'on'
            if sin_acceso:
                v1 = v2 = v3 = v4 = ""
                obs = "Sin acceso"
            else:
                v1 = request.form.get(f'v1_{planta}_{i}', '').strip()
                v2 = request.form.get(f'v2_{planta}_{i}', '').strip()
                v3 = request.form.get(f'v3_{planta}_{i}', '').strip()
                v4 = request.form.get(f'v4_{planta}_{i}', '').strip()
                obs = ""

            if not sin_acceso:
                if not es_valido_temp(v1, permitir_fos=(planta == 'selectiva')):
                    return jsonify({"error": f"Valor 1 inválido en {b['tag']}"}), 400
                if not es_valido_temp(v2): return jsonify({"error": f"Valor 2 inválido en {b['tag']}"}), 400
                if not es_valido_temp(v3): return jsonify({"error": f"Valor 3 inválido en {b['tag']}"}), 400
                if not es_valido_temp(v4): return jsonify({"error": f"Valor 4 inválido en {b['tag']}"}), 400

            # Usar INSERT OR REPLACE para sobrescribir o insertar
            conn.execute('''
                INSERT OR REPLACE INTO registros 
                (planta, tag, ubicacion, fecha, v1, v2, v3, v4, observacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (planta, b['tag'], b['ubicacion'], fecha, 
                  v1 or None, v2 or None, v3 or None, v4 or None, obs or None))

        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/eliminar_fecha', methods=['POST'])
def eliminar_fecha():
    """Eliminar todos los registros de una fecha específica."""
    dia = request.form.get('dia', '').strip()
    mes = request.form.get('mes', '').strip()
    ano = request.form.get('ano', '').strip()
    if not all([dia, mes, ano]):
        return jsonify({"error": "Fecha incompleta"}), 400

    fecha_str = f"{dia}/{mes}/{ano}"

    try:
        conn = sqlite3.connect('instance/temperaturas.db')
        cur = conn.cursor()
        cur.execute("DELETE FROM registros WHERE fecha = ?", (fecha_str,))
        filas_afectadas = cur.rowcount
        conn.commit()
        conn.close()

        # Determinar si la fecha eliminada era la fecha de hoy
        hoy_str = datetime.now().strftime('%d/%m/%y')
        era_fecha_de_hoy = (fecha_str == hoy_str)

        mensaje = f"Se eliminaron {filas_afectadas} registros para la fecha {fecha_str}."
        if era_fecha_de_hoy:
            mensaje += " La fecha de hoy se mantiene en la lista."
        else:
            mensaje += " La fecha ha sido removida de la lista."

        return jsonify({
            "success": True,
            "message": mensaje,
            "era_fecha_de_hoy": era_fecha_de_hoy,
            "fecha_eliminada": fecha_str
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/nuevo_registro', methods=['POST'])
def nuevo_registro():
    """Preparar para un nuevo registro agregando la fecha actual a la lista si no está."""
    fecha_hoy_str = datetime.now().strftime('%d/%m/%y')
    
    try:
        conn = sqlite3.connect('instance/temperaturas.db')
        cur = conn.cursor()
        # Verificar si la fecha ya existe en la BD (aunque sea con 0 registros)
        cur.execute("SELECT COUNT(*) FROM registros WHERE fecha = ?", (fecha_hoy_str,))
        count = cur.fetchone()[0]
        conn.close()
        
        # Si no existe, no necesitamos hacer nada más en la BD.
        # La fecha se agregará al selector en el frontend.
        
        return jsonify({
            "success": True,
            "fecha_nueva": fecha_hoy_str,
            "existe_en_bd": count > 0
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/exportar', methods=['POST'])
def exportar():
    """Exportar datos a Excel"""
    planta = request.form.get('planta')
    dia = request.form.get('dia', '').strip()
    mes = request.form.get('mes', '').strip()
    ano = request.form.get('ano', '').strip()
    if not all([dia, mes, ano]) or planta not in BOMBAS:
        return jsonify({"error": "Fecha o planta inválida"}), 400

    # --- CORRECCIÓN PARA EL AÑO ---
    # Tomar los dos últimos dígitos del año como cadena
    ano_2_digitos = ano.zfill(4)[-2:] # Asegura 4 dígitos y toma los últimos 2
    # Formato de fecha: dd-mm-yy
    fecha_excel = f"{dia}-{mes}-{ano_2_digitos}"
    # --- FIN CORRECCIÓN ---

    # Usar el archivo predeterminado para la planta
    archivo = ARCHIVO_COLECTIVA if planta == 'colectiva' else ARCHIVO_SELECTIVA

    try:
        if os.path.exists(archivo):
            wb = load_workbook(archivo)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        for i, b in enumerate(BOMBAS[planta]):
            tag = b['tag']
            sin_acceso = request.form.get(f'sin_acceso_{planta}_{i}', '') == 'on'
            
            if sin_acceso:
                v1 = v2 = v3 = v4 = ""
                observacion = "Sin acceso"
            else:
                v1 = request.form.get(f'v1_{planta}_{i}', '').strip()
                v2 = request.form.get(f'v2_{planta}_{i}', '').strip()
                v3 = request.form.get(f'v3_{planta}_{i}', '').strip()
                v4 = request.form.get(f'v4_{planta}_{i}', '').strip()
                observacion = ""

            if not sin_acceso:
                if not es_valido_temp(v1, permitir_fos=(planta == 'selectiva')):
                    return jsonify({"error": f"Valor 1 inválido en {tag}"}), 400
                if not es_valido_temp(v2): return jsonify({"error": f"Valor 2 inválido en {tag}"}), 400
                if not es_valido_temp(v3): return jsonify({"error": f"Valor 3 inválido en {tag}"}), 400
                if not es_valido_temp(v4): return jsonify({"error": f"Valor 4 inválido en {tag}"}), 400

            if not sin_acceso:
                v1 = normalizar_numero(v1) or ""
                v2 = normalizar_numero(v2) or ""
                v3 = normalizar_numero(v3) or ""
                v4 = normalizar_numero(v4) or ""

            hoja_objetivo = None
            if tag in wb.sheetnames:
                hoja_objetivo = tag
            else:
                hoja_cercana = encontrar_hoja_mas_parecida(tag, wb.sheetnames)
                if hoja_cercana:
                    hoja_objetivo = hoja_cercana
                else:
                    hoja_objetivo = tag
                    ws = wb.create_sheet(title=tag)
                    if planta == 'colectiva':
                        encabezado = ["TAG", "Ubicación", "Fecha", "Lado Bomba", "Lado Polea", "Lado bomba", "Lado Polea", "Sin acceso"]
                    else:
                        encabezado = ["TAG", "Ubicación", "Fecha", "Lado Libre", "Lado Polea", "Lado bomba", "Lado Polea", "Sin acceso"]
                    # Escribir encabezado en la fila 1 (asumiendo que no hay celdas combinadas allí)
                    for col, valor in enumerate(encabezado, start=2):
                        cell = ws.cell(row=1, column=col, value=valor)
                        cell.border = borde_sutil  # Aplicar borde sutil

            ws = wb[hoja_objetivo]

            # Buscar la fila exacta con el TAG y la FECHA a partir de la fila 4
            fila_a_modificar = None
            # Iterar solo sobre las filas que tienen datos en la columna TAG (B), desde la fila 4
            for row_num in range(4, ws.max_row + 1): # Comenzar desde la fila 4
                tag_celda = ws.cell(row=row_num, column=2).value # Columna B es índice 2
                fecha_celda = ws.cell(row=row_num, column=4).value # Columna D es índice 4 (Fecha)
                if tag_celda == tag and fecha_celda == fecha_excel:
                    fila_a_modificar = row_num
                    break # Encontramos la fila exacta, salir del bucle

            datos_fila = [tag, b['ubicacion'], fecha_excel, v1 or "", v2 or "", v3 or "", v4 or "", observacion]
            
            if fila_a_modificar:
                # Si se encontró una fila con el mismo TAG y FECHA, sobrescribir
                print(f"Sobrescribiendo fila {fila_a_modificar} para {tag} en {fecha_excel}")
                for col, valor in enumerate(datos_fila, start=2): # Comenzar en columna B (índice 2)
                    cell = ws.cell(row=fila_a_modificar, column=col, value=valor)
                    # Verificar si la celda es una celda fusionada
                    if cell.coordinate in ws.merged_cells:
                        print(f"Advertencia: No se puede escribir en la celda fusionada {cell.coordinate}. Omitiendo.")
                        continue # Saltar esta celda
                    cell.border = borde_sutil  # Aplicar borde sutil
                    if col <= 3:  # TAG y Ubicación
                        cell.alignment = Alignment(horizontal='left')
                    else:  # Fecha, temperaturas, observación
                        cell.alignment = Alignment(horizontal='right')
            else:
                # Si no se encontró una fila con el mismo TAG y FECHA, añadir una nueva fila
                print(f"Añadiendo nueva fila para {tag} en {fecha_excel}")
                # Encontrar la primera fila vacía a partir de la fila 4
                fila_nueva = 4
                while fila_nueva <= ws.max_row:
                    if ws.cell(row=fila_nueva, column=2).value is None and ws.cell(row=fila_nueva, column=4).value is None:
                        # Encontramos una fila vacía
                        break
                    fila_nueva += 1
                # Si todas las filas estaban ocupadas, añadir al final
                if fila_nueva > ws.max_row:
                    fila_nueva = ws.max_row + 1

                for col, valor in enumerate(datos_fila, start=2):
                    cell = ws.cell(row=fila_nueva, column=col, value=valor)
                    # Verificar si la celda es una celda fusionada
                    if cell.coordinate in ws.merged_cells:
                        print(f"Advertencia: No se puede escribir en la celda fusionada {cell.coordinate}. Omitiendo.")
                        continue # Saltar esta celda
                    cell.border = borde_sutil  # Aplicar borde sutil
                    if col <= 3:  # TAG y Ubicación
                        cell.alignment = Alignment(horizontal='left')
                    else:  # Fecha, temperaturas, observación
                        cell.alignment = Alignment(horizontal='right')

        wb.save(archivo)
        # Devolver el nombre del archivo para que el frontend lo use para descargar
        return jsonify({"success": True, "archivo": archivo})
    except Exception as e:
        return jsonify({"error": f"Error al exportar: {str(e)}"}), 500

# Nueva ruta para descargar el archivo exportado
@app.route('/descargar/<path:nombre_archivo>')
def descargar(nombre_archivo):
    """Descarga el archivo exportado."""
    # Verificar que el nombre del archivo no intente navegar fuera del directorio
    if '..' in nombre_archivo or nombre_archivo.startswith('/'):
        return jsonify({"error": "Nombre de archivo inválido"}), 400
    try:
        return send_file(nombre_archivo, as_attachment=True)
    except FileNotFoundError:
        return jsonify({"error": "Archivo no encontrado"}), 404
    except Exception as e:
        return jsonify({"error": f"Error al descargar: {str(e)}"}), 500

# Añadir esta ruta en tu app.py si no la tienes
@app.route('/estadisticas_datos')
def estadisticas_datos():
    """Devuelve datos históricos de temperaturas para una bomba específica."""
    planta = request.args.get('planta')
    tag = request.args.get('tag')
    
    if not planta or not tag or planta not in BOMBAS:
        return jsonify({"error": "Parámetros inválidos"}), 400

    # Verificar que el tag pertenezca a la planta
    tags_validos = [b['tag'] for b in BOMBAS[planta]]
    if tag not in tags_validos:
        return jsonify({"error": "TAG no válido para la planta"}), 400

    try:
        conn = sqlite3.connect('instance/temperaturas.db')
        cur = conn.cursor()
        # Seleccionar registros ordenados por fecha (convertimos a ISO para ordenar)
        cur.execute("""
            SELECT fecha, v1, v2, v3, v4 
            FROM registros 
            WHERE planta = ? AND tag = ? AND (v1 != '' OR v2 != '' OR v3 != '' OR v4 != '')
            ORDER BY 
                substr(fecha, 7, 2) || '-' || 
                substr(fecha, 4, 2) || '-' || 
                substr(fecha, 1, 2) ASC
        """, (planta, tag))
        filas = cur.fetchall()
        conn.close()

        # Convertir a lista de diccionarios
        datos = []
        for fecha, v1, v2, v3, v4 in filas:
            punto = {"fecha": fecha}
            # Convertir valores a número si es posible
            try:
                punto["v1"] = float(v1.replace(',', '.')) if v1 and v1 != "Fuera de servicio" else None
            except:
                punto["v1"] = None
            try:
                punto["v2"] = float(v2.replace(',', '.')) if v2 else None
            except:
                punto["v2"] = None
            try:
                punto["v3"] = float(v3.replace(',', '.')) if v3 else None
            except:
                punto["v3"] = None
            try:
                punto["v4"] = float(v4.replace(',', '.')) if v4 else None
            except:
                punto["v4"] = None
            datos.append(punto)

        return jsonify(datos)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# --- NUEVA RUTA: Modificar Fecha ---
@app.route('/modificar_fecha', methods=['POST'])
def modificar_fecha():
    """Modifica la fecha de todos los registros de una fecha específica a una nueva fecha."""
    try:
        # Obtener los datos del formulario
        dia_actual = request.form.get('dia_actual', '').strip()
        mes_actual = request.form.get('mes_actual', '').strip()
        ano_actual = request.form.get('ano_actual', '').strip()
        dia_nuevo = request.form.get('dia_nuevo', '').strip()
        mes_nuevo = request.form.get('mes_nuevo', '').strip()
        ano_nuevo = request.form.get('ano_nuevo', '').strip()

        # Validar que todos los campos existan y tengan 2 dígitos
        if not all([dia_actual, mes_actual, ano_actual, dia_nuevo, mes_nuevo, ano_nuevo]):
             return jsonify({
                'success': False,
                'error': 'Faltan parámetros requeridos.'
            }), 400

        # Opcional: Validar formato numérico simple (01-31, 01-12, 00-99)
        for p in [dia_actual, mes_actual, ano_actual, dia_nuevo, mes_nuevo, ano_nuevo]:
            if not p.isdigit() or len(p) != 2:
                 return jsonify({
                    'success': False,
                    'error': 'Formato de fecha inválido. Use dd/mm/yy (ej: 01/01/24).'
                }), 400

        fecha_actual = f"{dia_actual}/{mes_actual}/{ano_actual}"
        fecha_nueva = f"{dia_nuevo}/{mes_nuevo}/{ano_nuevo}"

        # Verificar si la nueva fecha ya existe en la base de datos
        conn = sqlite3.connect('instance/temperaturas.db')
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM registros WHERE fecha = ?", (fecha_nueva,))
        count_nueva = cur.fetchone()[0]
        if count_nueva > 0:
             conn.close()
             return jsonify({
                'success': False,
                'error': f'Ya existen registros para la nueva fecha: {fecha_nueva}.'
            }), 400

        # Verificar si existen registros con la fecha actual
        cur.execute("SELECT COUNT(*) FROM registros WHERE fecha = ?", (fecha_actual,))
        count_actual = cur.fetchone()[0]
        if count_actual == 0:
             conn.close()
             return jsonify({
                'success': False,
                'error': f'No se encontraron registros para la fecha actual: {fecha_actual}.'
            }), 404

        # Actualizar la fecha en la base de datos
        cur.execute("""
            UPDATE registros
            SET fecha = ?
            WHERE fecha = ?
        """, (fecha_nueva, fecha_actual))

        conn.commit()
        conn.close()

        # Devolver respuesta exitosa en JSON
        return jsonify({
            'success': True,
            'message': f'Fecha modificada de {fecha_actual} a {fecha_nueva}.',
            'fecha_anterior': fecha_actual,
            'fecha_nueva': fecha_nueva
        })

    except Exception as e:
        print(f"Error en /modificar_fecha: {e}") # Log para debugging
        return jsonify({
            'success': False,
            'error': 'Error interno del servidor al modificar la fecha.'
        }), 500
# --- FIN NUEVA RUTA ---

if __name__ == '__main__':
  
        app.run(debug=True, host='127.0.0.1', port=5000) # script completo funcional boton modificar fecha oscurece bordes y temperaturas