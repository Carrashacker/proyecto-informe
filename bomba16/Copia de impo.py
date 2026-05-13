import os
import sqlite3
from openpyxl import load_workbook
from datetime import datetime, timedelta

# --- CONFIGURACIÓN ---
archivo_excel_colectiva = "Control de temperaturas bombas colectivas.xlsx"
archivo_excel_selectiva = "Control de temperatura bombas selectiva.xlsx"
archivo_db = "instance/temperaturas.db"

# --- FUNCIONES DE APOYO ---
def parsear_fecha_excel(fecha_obj):
    """
    Intenta parsear una fecha desde el formato DD/MM/YY, DD/MM/YYYY, MM/DD/YY, MM/DD/YYYY o datetime del Excel.
    Devuelve un objeto datetime o None si falla.
    """
    if not fecha_obj:
        return None

    if isinstance(fecha_obj, datetime):
        return fecha_obj

    if isinstance(fecha_obj, str):
        fecha_str = fecha_obj
        # Intenta varios formatos comunes
        formatos = [
            '%d/%m/%y', '%d/%m/%Y', '%m/%d/%y', '%m/%d/%Y',
            '%d-%m-%y', '%d-%m-%Y', '%m-%d-%y', '%m-%d-%Y'
        ]
        for fmt in formatos:
            try:
                return datetime.strptime(fecha_str, fmt)
            except ValueError:
                continue
        # Si todos los formatos fallan, imprime un mensaje y devuelve None
        # print(f"  ⚠️  Formato de fecha no reconocido: '{fecha_str}'") # Opcional: comentar para menos verbosidad
        return None

    # print(f"  ⚠️  Tipo de dato de fecha no manejable: {type(fecha_obj)} - {fecha_obj}") # Opcional: comentar para menos verbosidad
    return None

def normalizar_numero(valor):
    """Convierte coma decimal a punto y valida rango. Devuelve string o None."""
    if valor == "": return ""
    valor = str(valor).replace(',', '.')
    # Manejar textos especiales que pueden aparecer en estos archivos
    # Convertimos a minúsculas para comparar
    valor_lower = valor.lower()
    if valor_lower in ["fuera de servicio", "sin acceso", "sin acceso", "planta detenida", "flujo detenido por operaciones", "bomba detenida, bx sin carga"]:
        # Devolver el valor original como string para la base de datos
        return valor
    # Manejar casos como "/" o "-" que no son números
    if valor_lower in ["/", "-", "n/a", "na"]:
         return ""
    try:
        f = float(valor)
        if 0 <= f <= 100:
            return str(f) # Devuelve como string para la base de datos
        else:
            print(f"  ⚠️  Valor fuera de rango [0-100]: {valor}")
            return None
    except ValueError:
        # Si no es un número y no es un texto especial, lo devuelve como string (por ejemplo, "Fuera de servicio")
        return str(valor)

def leer_datos_hoja_individual(nombre_archivo, nombre_hoja, planta):
    """
    Lee una hoja individual de un archivo Excel.
    Asume que los datos comienzan en la fila 4.
    Estructura: B=TAG, C=Ubicacion, D=Fecha, E=V1, F=V2, G=V3, H=V4, I=Obs (opcional).
    Devuelve una lista de datos y un conjunto de fechas únicas encontradas en esa hoja.
    """
    print(f"    📄 Procesando hoja: '{nombre_hoja}'")
    try:
        wb = load_workbook(nombre_archivo, read_only=True, data_only=True)
    except Exception as e:
        print(f"    ❌ Error al abrir '{nombre_archivo}': {e}")
        return [], set()

    if nombre_hoja not in wb.sheetnames:
        print(f"    ❌ Hoja '{nombre_hoja}' no encontrada en '{nombre_archivo}'")
        wb.close()
        return [], set()

    hoja = wb[nombre_hoja]
    datos = []
    fechas_set = set()

    # --- ESTRUCTURA FIJA DE COLUMNAS (basada en fila 4) ---
    idx_tag = 2      # Columna B
    idx_ubicacion = 3 # Columna C
    idx_fecha = 4    # Columna D
    idx_v1 = 5       # Columna E
    idx_v2 = 6       # Columna F
    idx_v3 = 7       # Columna G
    idx_v4 = 8       # Columna H
    idx_obs = 9      # Columna I (puede no existir)
    # --- FIN ESTRUCTURA FIJA ---

    # Procesar filas de datos (empezando desde la fila 4)
    for fila_num, fila in enumerate(hoja.iter_rows(min_row=4, values_only=True), 4):
        # Verificar si hay un TAG en la columna correspondiente (Columna B)
        if fila and fila[idx_tag - 1]: # Debe haber TAG (índices de openpyxl empiezan en 1, por eso -1)
            try:
                tag = str(fila[idx_tag - 1]).strip() # Índices de fila en Python empiezan en 0
                ubicacion = str(fila[idx_ubicacion - 1]).strip() if fila[idx_ubicacion - 1] else ""
                fecha_obj = fila[idx_fecha - 1]
                v1 = str(fila[idx_v1 - 1]).strip() if fila[idx_v1 - 1] is not None else ""
                v2 = str(fila[idx_v2 - 1]).strip() if fila[idx_v2 - 1] is not None else ""
                v3 = str(fila[idx_v3 - 1]).strip() if fila[idx_v3 - 1] is not None else ""
                v4 = str(fila[idx_v4 - 1]).strip() if fila[idx_v4 - 1] is not None else ""
                # Verificar si la columna de observación existe en esta fila
                obs = str(fila[idx_obs - 1]).strip() if len(fila) >= idx_obs and fila[idx_obs - 1] is not None else ""

                fecha_dt = parsear_fecha_excel(fecha_obj)
                if not fecha_dt:
                    print(f"      ⚠️  Fecha no válida en fila {fila_num}, TAG {tag}: {fecha_obj}")
                    continue # Saltar esta fila si la fecha no es válida

                # Añadir fecha al conjunto
                fechas_set.add(fecha_dt)
                # Añadir datos a la lista temporal, manteniendo el orden original
                datos.append({
                    'planta': planta,
                    'tag': tag,
                    'ubicacion': ubicacion,
                    'fecha_obj': fecha_dt,
                    'fecha_str': str(fecha_obj), # Guardamos el original para mostrarlo si es necesario
                    'v1': v1,
                    'v2': v2,
                    'v3': v3,
                    'v4': v4,
                    'obs': obs,
                    'hoja': nombre_hoja,
                    'fila_num': fila_num # Para mantener el orden original
                })
            except (IndexError, AttributeError) as e:
                print(f"      ⚠️  Error al leer fila {fila_num} en hoja '{nombre_hoja}': {e}")
                continue
        elif fila and not any(fila): # Fila vacía, fin de datos
             break # Salir del bucle de filas si la fila está vacía

    wb.close()
    return datos, fechas_set


# --- SCRIPT PRINCIPAL ---
if __name__ == "__main__":
    if not os.path.exists(archivo_excel_colectiva):
        print(f"❌ Error: Archivo Excel Colectiva '{archivo_excel_colectiva}' no encontrado.")
        exit(1)

    if not os.path.exists(archivo_excel_selectiva):
        print(f"❌ Error: Archivo Excel Selectiva '{archivo_excel_selectiva}' no encontrado.")
        exit(1)

    if not os.path.exists(archivo_db):
        print(f"❌ Error: Archivo de base de datos '{archivo_db}' no encontrado.")
        exit(1)

    print("📅 Leyendo fechas de ambos archivos (procesando hoja por hoja, fila 4)...")

    # Leer datos y fechas de cada hoja en ambos archivos
    datos_totales = []
    fechas_totales = set()

    print(f"🔄 Abriendo archivo: '{archivo_excel_colectiva}'")
    try:
        wb_colectiva = load_workbook(archivo_excel_colectiva, read_only=True)
        hojas_colectiva = wb_colectiva.sheetnames
        wb_colectiva.close()
    except Exception as e:
        print(f"❌ Error al leer las hojas de '{archivo_excel_colectiva}': {e}")
        exit(1)

    for nombre_hoja in hojas_colectiva:
        # Intentar leer la hoja como una hoja de datos
        datos_hoja, fechas_hoja = leer_datos_hoja_individual(archivo_excel_colectiva, nombre_hoja, 'colectiva')
        datos_totales.extend(datos_hoja)
        fechas_totales.update(fechas_hoja)

    print(f"🔄 Abriendo archivo: '{archivo_excel_selectiva}'")
    try:
        wb_selectiva = load_workbook(archivo_excel_selectiva, read_only=True)
        hojas_selectiva = wb_selectiva.sheetnames
        wb_selectiva.close()
    except Exception as e:
        print(f"❌ Error al leer las hojas de '{archivo_excel_selectiva}': {e}")
        exit(1)

    for nombre_hoja in hojas_selectiva:
        # Intentar leer la hoja como una hoja de datos
        datos_hoja, fechas_hoja = leer_datos_hoja_individual(archivo_excel_selectiva, nombre_hoja, 'selectiva')
        datos_totales.extend(datos_hoja)
        fechas_totales.update(fechas_hoja)


    # Combinar todas las fechas y obtener las 3 más recientes
    fechas_ordenadas = sorted(fechas_totales, reverse=True)
    fechas_a_usar = set(fechas_ordenadas[:3]) # Tomar las 3 más recientes

    if not fechas_a_usar:
        print("❌ No se encontraron fechas válidas en los archivos.")
        exit(1)

    print(f"📅 Fechas seleccionadas (últimas 3): {', '.join([f.strftime('%d/%m/%Y') for f in sorted(fechas_a_usar, reverse=True)])}")

    # Filtrar datos para las fechas seleccionadas
    datos_filtrados = []
    for dato in datos_totales:
        if dato['fecha_obj'] in fechas_a_usar:
            datos_filtrados.append(dato)

    # Ordenar los datos filtrados por planta (colectiva primero), hoja y luego por el número de fila original
    datos_filtrados.sort(key=lambda x: (x['planta'], x['hoja'], x['fila_num']))

    print(f"📥 Preparados {len(datos_filtrados)} registros de las últimas 3 fechas para subir...")

    # Conectar a la base de datos
    try:
        conn = sqlite3.connect(archivo_db)
        cursor = conn.cursor()
    except Exception as e:
        print(f"❌ Error al conectar a la base de datos: {e}")
        exit(1)

    registros_subidos = 0
    for dato in datos_filtrados:
        # Normalizar temperaturas
        v1_norm = normalizar_numero(dato['v1'])
        v2_norm = normalizar_numero(dato['v2'])
        v3_norm = normalizar_numero(dato['v3'])
        v4_norm = normalizar_numero(dato['v4'])

        # Convertir fecha al formato dd/mm/yy
        fecha_db = dato['fecha_obj'].strftime('%d/%m/%y')

        try:
            cursor.execute('''
                INSERT OR REPLACE INTO registros (planta, tag, ubicacion, fecha, v1, v2, v3, v4, observacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (dato['planta'], dato['tag'], dato['ubicacion'], fecha_db, v1_norm, v2_norm, v3_norm, v4_norm, dato['obs']))
            registros_subidos += 1
            print(f"  + {dato['tag']} ({fecha_db}) - {dato['planta']} - {dato['hoja']}")
        except sqlite3.Error as e:
            print(f"  ❌ Error al insertar {dato['tag']} ({fecha_db}): {e}")

    try:
        conn.commit()
        print(f"🎉 Se subieron {registros_subidos} registros exitosamente a '{archivo_db}'.")
    except sqlite3.Error as e:
        print(f"❌ Error al confirmar los cambios en la base de datos: {e}")
        conn.rollback()
    finally:
        conn.close()
